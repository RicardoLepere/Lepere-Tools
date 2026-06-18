"""
Logica de conversion Excel -> CSV.
Une todas las hojas de un archivo Excel en un unico CSV (UTF-8 con BOM).
"""

import csv
import os


def _normalize_header(value):
    if value is None:
        return ""
    return str(value).strip().lower()


def _row_is_empty(row):
    for cell in row:
        if cell is None:
            continue
        if isinstance(cell, str):
            if cell.strip() != "":
                return False
        else:
            return False
    return True


def _cell_to_text(value):
    if value is None:
        return ""
    return value if isinstance(value, str) else str(value)


def contar_hojas_y_tamano(ruta_entrada):
    """Devuelve (numero_de_hojas, tamano_en_bytes) sin cargar todo el archivo."""
    tamano = os.path.getsize(ruta_entrada)
    ext = os.path.splitext(ruta_entrada)[1].lower()
    if ext == ".xlsx":
        from openpyxl import load_workbook

        wb = load_workbook(ruta_entrada, read_only=True, data_only=True)
        try:
            n = len(wb.sheetnames)
        finally:
            wb.close()
        return n, tamano
    elif ext == ".xls":
        import xlrd

        book = xlrd.open_workbook(ruta_entrada, on_demand=True)
        try:
            n = book.nsheets
        finally:
            book.release_resources()
        return n, tamano
    else:
        raise ValueError(f"Extension no soportada: {ext}")


def convertir_xlsx(ruta_entrada, ruta_salida, progreso_cb):
    """Convierte un .xlsx usando openpyxl en modo read_only (streaming)."""
    from openpyxl import load_workbook

    wb = load_workbook(ruta_entrada, read_only=True, data_only=True)
    try:
        nombres_hojas = wb.sheetnames
        total_hojas = len(nombres_hojas)
        encabezados_ref = None

        with open(ruta_salida, "w", encoding="utf-8-sig", newline="") as f:
            writer = csv.writer(f)

            for idx_hoja, nombre in enumerate(nombres_hojas, start=1):
                ws = wb[nombre]
                progreso_cb(
                    fase="hoja",
                    hoja_actual=idx_hoja,
                    total_hojas=total_hojas,
                    nombre_hoja=nombre,
                    filas=0,
                )

                filas_escritas = 0
                primera_fila = True

                for fila in ws.iter_rows(values_only=True):
                    if primera_fila:
                        primera_fila = False
                        if encabezados_ref is None:
                            encabezados_ref = [_normalize_header(c) for c in fila]
                            writer.writerow([_cell_to_text(c) for c in fila])
                            filas_escritas += 1
                            continue
                        actual = [_normalize_header(c) for c in fila]
                        if actual == encabezados_ref:
                            continue
                        if _row_is_empty(fila):
                            continue
                        writer.writerow([_cell_to_text(c) for c in fila])
                        filas_escritas += 1
                        continue

                    if _row_is_empty(fila):
                        continue
                    writer.writerow([_cell_to_text(c) for c in fila])
                    filas_escritas += 1

                    if filas_escritas % 1000 == 0:
                        progreso_cb(
                            fase="filas",
                            hoja_actual=idx_hoja,
                            total_hojas=total_hojas,
                            nombre_hoja=nombre,
                            filas=filas_escritas,
                        )

                progreso_cb(
                    fase="hoja_fin",
                    hoja_actual=idx_hoja,
                    total_hojas=total_hojas,
                    nombre_hoja=nombre,
                    filas=filas_escritas,
                )

                try:
                    del wb[nombre]
                except Exception:
                    pass
    finally:
        wb.close()


def convertir_xls(ruta_entrada, ruta_salida, progreso_cb):
    """Convierte un .xls clasico usando xlrd."""
    import xlrd

    book = xlrd.open_workbook(ruta_entrada, on_demand=True)
    try:
        total_hojas = book.nsheets
        encabezados_ref = None

        with open(ruta_salida, "w", encoding="utf-8-sig", newline="") as f:
            writer = csv.writer(f)

            for idx_hoja in range(total_hojas):
                sheet = book.sheet_by_index(idx_hoja)
                nombre = sheet.name
                progreso_cb(
                    fase="hoja",
                    hoja_actual=idx_hoja + 1,
                    total_hojas=total_hojas,
                    nombre_hoja=nombre,
                    filas=0,
                )

                filas_escritas = 0
                for r in range(sheet.nrows):
                    fila = sheet.row_values(r)

                    if r == 0:
                        if encabezados_ref is None:
                            encabezados_ref = [_normalize_header(c) for c in fila]
                            writer.writerow([_cell_to_text(c) for c in fila])
                            filas_escritas += 1
                            continue
                        actual = [_normalize_header(c) for c in fila]
                        if actual == encabezados_ref:
                            continue
                        if _row_is_empty(fila):
                            continue
                        writer.writerow([_cell_to_text(c) for c in fila])
                        filas_escritas += 1
                        continue

                    if _row_is_empty(fila):
                        continue
                    writer.writerow([_cell_to_text(c) for c in fila])
                    filas_escritas += 1

                    if filas_escritas % 1000 == 0:
                        progreso_cb(
                            fase="filas",
                            hoja_actual=idx_hoja + 1,
                            total_hojas=total_hojas,
                            nombre_hoja=nombre,
                            filas=filas_escritas,
                        )

                progreso_cb(
                    fase="hoja_fin",
                    hoja_actual=idx_hoja + 1,
                    total_hojas=total_hojas,
                    nombre_hoja=nombre,
                    filas=filas_escritas,
                )
                book.unload_sheet(idx_hoja)
    finally:
        book.release_resources()


def convertir(ruta_entrada, progreso_cb):
    carpeta = os.path.dirname(ruta_entrada)
    base = os.path.splitext(os.path.basename(ruta_entrada))[0]
    ruta_salida = os.path.join(carpeta, f"{base}_unificado.csv")

    ext = os.path.splitext(ruta_entrada)[1].lower()
    if ext == ".xlsx":
        convertir_xlsx(ruta_entrada, ruta_salida, progreso_cb)
    elif ext == ".xls":
        convertir_xls(ruta_entrada, ruta_salida, progreso_cb)
    else:
        raise ValueError(f"Extension no soportada: {ext}")

    return ruta_salida
